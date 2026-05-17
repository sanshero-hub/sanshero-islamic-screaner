"""
Excel export — professional formatted output with AAOIFI + Damodaran results
"""

import io, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY    = "1B3A5C"
WHITE   = "FFFFFF"
GREEN_BG = "E6F4EA"; GREEN_FG = "1E6B3C"
AMBER_BG = "FFF8E1"; AMBER_FG = "7A5200"
RED_BG   = "FDECEA"; RED_FG   = "9B1C1C"
ALT_ROW  = "F7F9FC"
META_BG  = "F0F4F8"
BORDER_C = "D0D7DE"

THIN   = Side(style="thin", color=BORDER_C)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Ticker",               10),
    ("Company Name",         28),
    ("Sector",               16),
    ("Industry",             20),
    ("Shariah Status",       14),
    ("Source",               12),
    ("Debt / MktCap",        13),
    ("Cash / MktCap",        13),
    ("Impure Income %",      14),
    ("Compliance Reason",    24),
    ("Current Price ($)",    14),
    ("52Wk High ($)",        12),
    ("52Wk Low ($)",         12),
    ("Intrinsic Value ($)",  15),
    ("Margin of Safety",     15),
    ("WACC",                  8),
    ("Terminal Growth",      13),
    ("P/E (Company)",        12),
    ("P/E (Sector Avg)",     13),
    ("EV/EBITDA (Co.)",      13),
    ("EV/EBITDA (Sector)",   14),
    ("Rel. Assessment",      14),
    ("Final Verdict",        12),
    ("Rationale",            40),
    ("Purification %",       13),
    ("Run Date",             12),
]


def _font(bold=False, color="000000", size=9, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _write_header(ws):
    for ci, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = _font(bold=True, color=WHITE, size=9)
        cell.fill      = _fill(NAVY)
        cell.alignment = _align("center", wrap=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col_width
    ws.row_dimensions[1].height = 30


def _pct(v):
    return f"{v:.1%}" if v is not None else ""


def _num(v, d=2):
    try:
        return round(float(v), d) if v is not None else ""
    except Exception:
        return ""


def _mul(v):
    try:
        return f"{float(v):.1f}x" if v is not None else ""
    except Exception:
        return ""


def _write_row(ws, row_idx, r, run_mode):
    alt    = row_idx % 2 == 0
    status = r.get("shariah_status", "")
    verdict= r.get("verdict", "")

    compliance_cols = run_mode in ("compliance", "both")
    valuation_cols  = run_mode in ("valuation", "both")

    mos_raw = r.get("margin_of_safety")
    try:
        mos_str = f"{float(mos_raw):.1f}%" if mos_raw is not None else ""
    except Exception:
        mos_str = ""

    row_data = [
        r.get("ticker", ""),
        r.get("company", ""),
        r.get("sector", ""),
        r.get("industry", ""),
        status            if compliance_cols else "—",
        r.get("shariah_source", "") if compliance_cols else "—",
        _pct(r.get("debt_ratio"))          if compliance_cols else "",
        _pct(r.get("cash_ratio"))          if compliance_cols else "",
        _pct(r.get("impure_income_ratio")) if compliance_cols else "",
        r.get("fail_reason", "")           if compliance_cols else "",
        _num(r.get("current_price")),
        _num(r.get("week52_high")),
        _num(r.get("week52_low")),
        _num(r.get("intrinsic_value"))     if valuation_cols else "",
        mos_str                            if valuation_cols else "",
        f"{r.get('wacc'):.1f}%"           if (valuation_cols and r.get("wacc")) else "",
        f"{r.get('terminal_growth'):.1f}%" if (valuation_cols and r.get("terminal_growth")) else "",
        _mul(r.get("pe_company"))          if valuation_cols else "",
        _mul(r.get("pe_sector"))           if valuation_cols else "",
        _mul(r.get("ev_ebitda_company"))   if valuation_cols else "",
        _mul(r.get("ev_ebitda_sector"))    if valuation_cols else "",
        r.get("relative_assessment", "")  if valuation_cols else "",
        verdict                            if valuation_cols else "",
        r.get("rationale", "")            if valuation_cols else "",
        f"{r.get('purification_pct'):.2f}%" if (valuation_cols and r.get("purification_pct")) else "",
        r.get("run_date", ""),
    ]

    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.font      = _font()
        cell.border    = BORDER
        cell.alignment = _align(wrap=(ci in (2, 10, 24)))
        if alt:
            cell.fill = _fill(ALT_ROW)

    # Shariah status colouring (col 5)
    sc = ws.cell(row=row_idx, column=5)
    if status == "HALAL":
        sc.fill = _fill(GREEN_BG); sc.font = _font(bold=True, color=GREEN_FG)
    elif status == "PURIFY":
        sc.fill = _fill(AMBER_BG); sc.font = _font(bold=True, color=AMBER_FG)
    elif status == "HARAM":
        sc.fill = _fill(RED_BG);   sc.font = _font(bold=True, color=RED_FG)

    # Verdict colouring (col 23)
    vc = ws.cell(row=row_idx, column=23)
    if verdict == "BUY":
        vc.fill = _fill(GREEN_BG); vc.font = _font(bold=True, color=GREEN_FG)
    elif verdict == "HOLD":
        vc.fill = _fill(AMBER_BG); vc.font = _font(bold=True, color=AMBER_FG)
    elif verdict == "AVOID":
        vc.fill = _fill(RED_BG);   vc.font = _font(bold=True, color=RED_FG)

    # MoS colouring (col 15)
    mc = ws.cell(row=row_idx, column=15)
    try:
        mv = float(str(mos_str).replace("%", ""))
        if mv >= 20:
            mc.font = _font(bold=True, color=GREEN_FG)
        elif mv < -10:
            mc.font = _font(color=RED_FG)
    except Exception:
        pass

    ws.row_dimensions[row_idx].height = 16


def _add_run_info_sheet(wb, config):
    ws = wb.create_sheet("Run Info", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 48

    rows = [
        ("Islamic Stock Screener", ""),
        ("", ""),
        ("Run Date",          config.get("run_date", "")),
        ("Run Mode",          config.get("run_mode", "")),
        ("Sectors Screened",  config.get("sectors", "")),
        ("Output Filter",     config.get("output_filter", "")),
        ("", ""),
        ("Shariah Standard",  "AAOIFI Standard No. 21"),
        ("Debt/MktCap Limit", "< 30%"),
        ("Cash/MktCap Limit", "< 30%"),
        ("Impure Income Limit","< 5% of revenue"),
        ("Valuation Method",  "Damodaran DCF + Relative Valuation"),
        ("BUY threshold",     "Margin of Safety ≥ 20%"),
        ("HOLD threshold",    "MoS between -10% and 20%"),
        ("Data Source",       "Financial Modeling Prep (FMP)"),
        ("Valuation Engine",  "Anthropic Claude (claude-opus-4-5)"),
        ("", ""),
        ("Total Screened",    config.get("total", 0)),
        ("HALAL",             config.get("halal", 0)),
        ("PURIFY",            config.get("purify", 0)),
        ("HARAM",             config.get("haram", 0)),
        ("BUY Signals",       config.get("buy", 0)),
        ("HOLD Signals",      config.get("hold", 0)),
        ("", ""),
        ("Disclaimer", "Educational use only — not financial advice."),
        ("",           "Verify all data before investment decisions."),
    ]

    for i, (k, v) in enumerate(rows, 1):
        ka = ws.cell(row=i, column=1, value=k)
        va = ws.cell(row=i, column=2, value=str(v) if v != "" else "")
        if i == 1:
            ka.font = Font(name="Arial", size=13, bold=True, color=NAVY)
        elif k and v != "":
            ka.font = _font(bold=True, size=9)
            va.font = _font(size=9)
            ka.fill = _fill(META_BG)
            va.fill = _fill(META_BG)
            ka.border = BORDER
            va.border = BORDER


def build_excel(results: list, config: dict, output_filter: str, run_mode: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Analysis"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    _write_header(ws)

    row_idx  = 2
    run_date = config.get("run_date", str(datetime.date.today()))

    counts = {"total": len(results), "halal": 0, "purify": 0,
              "haram": 0, "buy": 0, "hold": 0}

    for r in results:
        status  = r.get("shariah_status", "")
        verdict = r.get("verdict", "")

        if output_filter == "halal_only" and status not in ("HALAL", "PURIFY"):
            continue
        if output_filter == "buy_only" and verdict != "BUY":
            continue
        if output_filter == "buy_hold" and verdict not in ("BUY", "HOLD"):
            continue

        r["run_date"] = run_date
        _write_row(ws, row_idx, r, run_mode)
        row_idx += 1

        if status == "HALAL":   counts["halal"]  += 1
        elif status == "PURIFY":counts["purify"] += 1
        elif status == "HARAM": counts["haram"]  += 1
        if verdict == "BUY":    counts["buy"]    += 1
        elif verdict == "HOLD": counts["hold"]   += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    config.update(counts)
    _add_run_info_sheet(wb, config)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
